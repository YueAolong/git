from openpyxl import load_workbook

wb = load_workbook('热管理日报.xlsx')
ws = wb.active

# 打印每一行的数据
for row in ws.iter_rows(min_row=2, values_only=True):
    print(row)
