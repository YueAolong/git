from openpyxl import load_workbook

# 加载 Excel 文件
wb = load_workbook('热管理日报.xlsx')  # 把文件名替换成你的文件
ws = wb.active  # 获取当前活动工作表
ws.cell(row=1, column=5, value="是否超标")  # 表头

for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
    temp = row[2].value
    flag = "是" if temp > 40 else "否"
    ws.cell(row=i, column=5, value=flag)
wb.save('热管理日报_标记超标.xlsx')

