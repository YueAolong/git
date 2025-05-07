from openpyxl import load_workbook

# 加载 Excel 文件
wb = load_workbook('热管理日报.xlsx')  # 把文件名替换成你的文件
ws = wb.active  # 获取当前活动工作表

# 遍历数据
for row in ws.iter_rows(min_row=2):  # 从第2行开始（跳过表头）
    temp = row[2].value  # 第3列是温度（C列）
    if temp and temp > 40:
        print(f"高温警告: {row[0].value} 部位温度为 {temp}℃")

